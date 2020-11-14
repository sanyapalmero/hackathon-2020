from decimal import Decimal
from typing import Any, Dict, Iterator, List, Tuple

from openpyxl import load_workbook

from assets.models import Asset, XlsImportColumnMatch


class XlsImportError(Exception):
    """Ошибка импорта XLS. Содержит сообщение об ошибке, которое можно показать пользователю."""


class XlsAssetsFile:
    """
    Загруженный и распаршенный файл XLS с имуществом.
    """

    def __init__(self, file_obj) -> None:
        try:
            self.workbook = load_workbook(file_obj)
        except Exception:  # никому не говори
            raise XlsImportError("Загруженный файл должен быть в формате XLS или XLSX")

        self.sheet = self.workbook.worksheets[0]

    def count_columns(self):
        # max_column is 1-based
        return self.sheet.max_column

    def _is_row_empty(self, row):
        return all(c is None for c in row)

    def import_assets(
        self, skip_lines: int, column_matches: List[XlsImportColumnMatch]
    ) -> Iterator[Asset]:
        column_matches_dict = {}
        for match in column_matches:
            column_matches_dict[match.asset_attribute] = match

        for row in self.sheet.iter_rows(min_row=skip_lines + 1, values_only=True):
            if self._is_row_empty(row):
                continue

            asset = Asset()
            for attr_name, importer in ATTRIBUTE_IMPORTERS.items():
                if attr_name in column_matches_dict:
                    selected_column = column_matches_dict[attr_name].column_index
                else:
                    selected_column = None

                if selected_column:
                    try:
                        raw_value = row[selected_column] or ""
                    except IndexError:
                        raw_value = ""

                    value = importer.translate_value(raw_value)

                else:
                    value = importer.default

                if value is not None:
                    setattr(asset, attr_name, value)

            yield asset


class Importer:
    def __init__(self, default=None) -> None:
        self.default = default

    def translate_value(self, xls_value: Any) -> Any:
        raise NotImplementedError("translate_value")


class TextImporter(Importer):
    def translate_value(self, xls_value: Any) -> str:
        return str(xls_value)


class SubstringMatcher(Importer):
    def __init__(self, matches: Dict[str, Any], **kwargs) -> None:
        super().__init__(**kwargs)
        self.matches = matches

    def translate_value(self, xls_value: Any) -> str:
        xls_value = str(xls_value)

        for key, value in self.matches.items():
            if key in xls_value:
                return value

        return None


class DecimalImporter(Importer):
    def translate_value(self, xls_value: Any) -> str:
        return round(Decimal(xls_value), 1)


ATTRIBUTE_IMPORTERS = {
    "balance_holder": TextImporter(),
    "name": TextImporter(),
    "type_asset": SubstringMatcher(
        {
            "движимое": Asset.TypeAsset.MOVABLE,
            "недвижимое": Asset.TypeAsset.IMMOVABLE,
        },
        default=Asset.TypeAsset.IMMOVABLE,
    ),
    "full_name_contact_person": TextImporter(),
    "phone_contact_person": TextImporter(),
    "email_contact_person": TextImporter(),
    "address": TextImporter(),
    "square": DecimalImporter(),
    "cadastral_number": TextImporter(),
    "state": SubstringMatcher(
        {
            "ремонт": Asset.State.USABLE_WITH_REPAIR,
            "не пригоден": Asset.State.UNUSABLE,
            "": Asset.State.USABLE,
        },
        default=Asset.State.USABLE,
    ),
    "state_comment": TextImporter(),
}


def list_importable_attributes() -> List[Tuple[str, str]]:
    """
    Возвращает список тюплов вида ("атрибут", "человеко-читаемое название атрибута").
    """
    importable_attributes = []
    for attr_name in ATTRIBUTE_IMPORTERS:
        verbose_name = Asset._meta.get_field(attr_name).verbose_name
        importable_attributes.append((attr_name, verbose_name))
    return importable_attributes
