import json
import string
from tempfile import NamedTemporaryFile

from django.db import models
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls.base import reverse_lazy
from django.utils.decorators import method_decorator
from django.views import generic
from openpyxl import Workbook

from users.decorators import role_required
from users.models import User

from .forms import (
    AssetMailForm,
    AssetSearchForm,
    ExportXlsForm,
    ImmovableAssetForm,
    ImportXlsSelectFileForm,
    MovableAssetForm,
    ResolutionForm,
)
from .models import (
    Asset,
    AssetPhoto,
    KindAsset,
    Resolution,
    XlsImport,
    XlsImportColumnMatch,
)
from .services.xlsimport import XlsAssetsFile, list_importable_attributes


@method_decorator(role_required(User.ROLE_ADMIN, User.ROLE_USER), name="dispatch")
class AssetsListView(generic.View):
    mpr_template_name = "assets/mpr/assets_list.html"
    ogv_template_name = "assets/ogv/assets_list.html"

    search_form_class = AssetSearchForm

    def _search_filter(self, assets_qs, search_form):
        balance_holder = search_form.cleaned_data["balance_holder"]
        if balance_holder:
            assets_qs = assets_qs.filter(balance_holder__icontains=balance_holder)

        name = search_form.cleaned_data["name"]
        if name:
            assets_qs = assets_qs.filter(name__icontains=name)

        type_asset = search_form.cleaned_data["type_asset"]
        if type_asset:
            assets_qs = assets_qs.filter(type_asset=type_asset)

        expiration_date_start = search_form.cleaned_data["expiration_date_start"]
        if expiration_date_start:
            assets_qs = assets_qs.filter(expiration_date__gte=expiration_date_start)

        expiration_date_end = search_form.cleaned_data["expiration_date_end"]
        if expiration_date_start:
            assets_qs = assets_qs.filter(expiration_date__lte=expiration_date_end)

        address = search_form.cleaned_data["address"]
        if address:
            assets_qs = assets_qs.filter(address__icontains=address)

        square_start = search_form.cleaned_data["square_start"]
        if square_start:
            assets_qs = assets_qs.filter(square__gte=square_start)

        square_end = search_form.cleaned_data["square_end"]
        if square_end:
            assets_qs = assets_qs.filter(square__lte=square_end)

        cadastral_number = search_form.cleaned_data["cadastral_number"]
        if cadastral_number:
            assets_qs = assets_qs.filter(cadastral_number__icontains=cadastral_number)

        state = search_form.cleaned_data["state"]
        if state:
            assets_qs = assets_qs.filter(state=state)

        return assets_qs

    def _get_admin(self, request, kind_asset):
        assets_qs = None

        if kind_asset == KindAsset.NEW.value:
            assets_qs = Asset.objects.new_assets()
        if kind_asset == KindAsset.CONST.value:
            assets_qs = Asset.objects.cost_assets()
        if kind_asset == KindAsset.ARCHIVE.value:
            assets_qs = Asset.objects.archive_assets()
        if kind_asset == KindAsset.WITH_APPLICANTS.value:
            assets_qs = Asset.objects.with_applicants_assets()

        assets_dicts_list = [asset.get_asset_info() for asset in assets_qs]
        assets_json = json.dumps(assets_dicts_list, ensure_ascii=False)

        search_form = self.search_form_class(request.GET)
        if search_form.is_valid():
            assets_qs = self._search_filter(assets_qs, search_form)

        return render(
            request,
            self.mpr_template_name,
            context={
                "assets_qs": assets_qs,
                "kind_asset": kind_asset,
                "assets_json": assets_json,
                "search_form": search_form,
            },
        )

    def _get_user(self, request, kind_asset):
        assets_qs = None

        if kind_asset == KindAsset.ARCHIVE.value:
            raise Http404()

        if kind_asset == KindAsset.NEW.value:
            assets_qs = Asset.objects.new_assets()
        if kind_asset == KindAsset.CONST.value:
            assets_qs = Asset.objects.cost_assets()

        assets_dicts_list = [asset.get_asset_info() for asset in assets_qs]
        assets_json = json.dumps(assets_dicts_list, ensure_ascii=False)

        search_form = self.search_form_class(request.GET)
        if search_form.is_valid():
            assets_qs = self._search_filter(assets_qs, search_form)

        return render(
            request,
            self.ogv_template_name,
            context={
                "assets_qs": assets_qs,
                "kind_asset": kind_asset,
                "assets_json": assets_json,
                "search_form": search_form,
            },
        )

    def get(self, request, kind_asset):
        if kind_asset not in KindAsset:
            raise Http404()

        user = request.user

        if user.is_admin:
            return self._get_admin(request, kind_asset)
        elif user.is_user:
            return self._get_user(request, kind_asset)
        else:
            raise Http404()


@method_decorator(role_required(User.ROLE_ADMIN, User.ROLE_USER), name="dispatch")
class AssetDetailView(generic.DetailView):
    mpr_template_name = "assets/mpr/asset_detail.html"
    ogv_template_name = "assets/ogv/asset_detail.html"
    template_name = None

    def get(self, request, pk):
        user = request.user

        if user.is_admin:
            self.template_name = self.mpr_template_name
        if user.is_user:
            self.template_name = self.ogv_template_name

        asset = Asset.objects.get(pk=pk)
        photos = AssetPhoto.objects.filter(asset=asset)
        approved_resolutions = Resolution.objects.filter(
            asset=asset, kind=Resolution.Kind.APPROVED
        ).order_by("created_at")
        refused_resolutions = Resolution.objects.filter(
            asset=asset, kind=Resolution.Kind.REFUSED
        ).order_by("created_at")

        return render(
            request,
            self.template_name,
            context={
                "asset": asset,
                "photos": photos,
                "approved_resolutions": approved_resolutions,
                "refused_resolutions": refused_resolutions,
            },
        )


@method_decorator(role_required(User.ROLE_ADMIN), name="dispatch")
class ArchiveAssetView(generic.View):
    def post(self, request):
        asset_id = request.POST.get("asset_id")

        asset = Asset.objects.get(pk=asset_id)
        asset.status = Asset.Status.ARCHIVED
        asset.save()

        return redirect(request.POST.get("back", "/"))


@method_decorator(role_required(User.ROLE_ADMIN), name="dispatch")
class ConstAssetView(generic.View):
    def post(self, request):
        asset_id = request.POST.get("asset_id")

        asset = Asset.objects.get(pk=asset_id)
        asset.expiration_date = None
        asset.save()

        return redirect(request.POST.get("back", "/"))


@method_decorator(role_required(User.ROLE_ADMIN), name="dispatch")
class AssetCreateView(generic.View):
    template_name = "assets/asset_form.html"
    movable_form_class = MovableAssetForm
    immovable_form_class = ImmovableAssetForm

    def movable_form_validate(self, request):
        movable_form = self.movable_form_class(request.POST)
        immovable_form = self.immovable_form_class()

        if not movable_form.is_valid():
            return render(
                request,
                self.template_name,
                context={
                    "validation_called": True,
                    "movable_form": movable_form,
                    "immovable_form": immovable_form,
                },
            )

        return movable_form

    def immovable_form_validate(self, request):
        immovable_form = self.immovable_form_class(request.POST)
        movable_form = self.movable_form_class()

        if not immovable_form.is_valid():
            return render(
                request,
                self.template_name,
                context={
                    "validation_called": True,
                    "immovable_form_received": True,
                    "movable_form": movable_form,
                    "immovable_form": immovable_form,
                },
            )

        return immovable_form

    def get(self, request):
        movable_form = self.movable_form_class()
        immovable_form = self.immovable_form_class()

        return render(
            request,
            self.template_name,
            context={"movable_form": movable_form, "immovable_form": immovable_form},
        )

    def post(self, request):
        type_asset = request.POST.get("type_asset")

        if type_asset == Asset.TypeAsset.MOVABLE.value:
            result = self.movable_form_validate(request)

        if type_asset == Asset.TypeAsset.IMMOVABLE.value:
            result = self.immovable_form_validate(request)

        if isinstance(result, HttpResponse):
            return result

        asset = result.save(commit=False)
        asset.type_asset = type_asset
        asset.save()

        photos = request.FILES.getlist("photos")
        for photo in photos:
            AssetPhoto.objects.create(asset=asset, photo=photo)

        ogv_users = User.objects.filter(role=User.ROLE_USER)
        for user in ogv_users:
            user.send_email(
                subject="Новое объявление",
                templates_name="users/email/asset",
                context={
                    "id": asset.id,
                    "name": asset.name,
                    "link": self.request.build_absolute_uri(asset.get_absolute_url()),
                },
            )

        return redirect(asset)


@method_decorator(role_required(User.ROLE_USER), name="dispatch")
class AssetMailView(generic.View):
    form_class = AssetMailForm
    template_name = "assets/mail.html"
    success_url = reverse_lazy("assets:assets-list", kwargs={"kind_asset": "new"})

    def get(self, request):
        return render(
            request,
            self.template_name,
            {
                "form": self.form_class(),
            },
        )

    def post(self, request):
        form = self.form_class(request.POST)

        if not form.is_valid():
            return render(
                request,
                self.template_name,
                {
                    "form": form,
                    "validation_called": True,
                },
            )

        admin_users = User.objects.filter(role=User.ROLE_ADMIN)
        for user in admin_users:
            user.send_email(
                subject="Письмо о высвобождении имущества",
                templates_name="users/email/asset-request",
                context={
                    "data": form.cleaned_data,
                },
            )

        return redirect(self.success_url)


@method_decorator(role_required(User.ROLE_ADMIN), name="dispatch")
class AssetUpdateView(generic.View):
    template_name = "assets/asset_form_update.html"
    movable_form_class = MovableAssetForm
    immovable_form_class = ImmovableAssetForm

    def get(self, request, pk):
        asset = get_object_or_404(Asset, id=pk)

        if asset.type_asset == Asset.TypeAsset.MOVABLE.value:
            form = self.movable_form_class(instance=asset)

        if asset.type_asset == Asset.TypeAsset.IMMOVABLE.value:
            form = self.immovable_form_class(instance=asset)

        return render(
            request, self.template_name, context={"form": form, "asset": asset}
        )

    def post(self, request, pk):
        asset = get_object_or_404(Asset, id=pk)

        if asset.type_asset == Asset.TypeAsset.MOVABLE.value:
            form = self.movable_form_class(request.POST, instance=asset)

        if asset.type_asset == Asset.TypeAsset.IMMOVABLE.value:
            form = self.immovable_form_class(request.POST, instance=asset)

        if not form.is_valid():
            return render(
                request,
                self.template_name,
                context={"form": form, "asset": asset, "validation_called": True},
            )

        photos_qs = AssetPhoto.objects.filter(asset=asset)
        new_photos = request.FILES.getlist("photos")

        if photos_qs and new_photos:
            photos_qs.delete()

        for photo in new_photos:
            AssetPhoto.objects.create(asset=asset, photo=photo)

        asset = form.save()

        return redirect(asset)


class AssetProtocolView(generic.DetailView):
    template_name = "assets/protocol.html"
    model = Asset


@method_decorator(role_required(User.ROLE_USER), name="dispatch")
class RefusedAssetView(generic.View):
    def post(self, request, pk):
        asset = Asset.objects.get(pk=pk)
        resolution = Resolution(asset=asset, user=request.user)
        resolution.kind = Resolution.Kind.REFUSED
        resolution.save()

        return redirect(asset)


class ApprovedAssetView(generic.View):
    template_name = "assets/ogv/resolution_form.html"
    form_class = ResolutionForm

    def get(self, request, pk):
        form = self.form_class()
        asset = Asset.objects.get(pk=pk)
        return render(
            request, self.template_name, context={"form": form, "asset": asset}
        )

    def post(self, request, pk):
        asset = Asset.objects.get(pk=pk)
        form = self.form_class(request.POST)

        if not form.is_valid():
            return render(
                request,
                self.template_name,
                context={
                    "form": form,
                    "asset": asset,
                    "validation_called": True,
                },
            )

        resolution = form.save(commit=False)
        resolution.user = request.user
        resolution.asset = asset
        resolution.kind = Resolution.Kind.APPROVED
        resolution.save()

        admin_users = User.objects.filter(role=User.ROLE_ADMIN)
        for user in admin_users:
            user.send_email(
                subject="Новое согласие",
                templates_name="users/email/asset",
                context={
                    "id": asset.id,
                    "name": asset.name,
                    "link": self.request.build_absolute_uri(asset.get_absolute_url()),
                },
            )

        return redirect(asset)


@method_decorator(role_required(User.ROLE_ADMIN), name="dispatch")
class ImportXlsSelectFileView(generic.FormView):
    template_name = "assets/import-xls/select-file.html"
    form_class = ImportXlsSelectFileForm

    def form_valid(self, form):
        xls_import = XlsImport.objects.create(
            file=form.cleaned_data["file"],
        )
        return redirect("assets:import-xls-match-columns", pk=xls_import.pk)


@method_decorator(role_required(User.ROLE_ADMIN), name="dispatch")
class ImportXlsMatchColumnsView(generic.View):
    template_name = "assets/import-xls/match-columns.html"

    def get(self, request, pk):
        xls_import = get_object_or_404(XlsImport, pk=pk)
        xls_asset_file = XlsAssetsFile(xls_import.file)

        column_choices = [(-1, "Нет")]
        for i in range(xls_asset_file.count_columns()):
            column_choices.append((i, string.ascii_uppercase[i]))

        attribute_contexts = []
        for attr_name, verbose_name in list_importable_attributes():
            try:
                selected_column = xls_import.column_matches.get(
                    asset_attribute=attr_name
                ).column_index
            except XlsImportColumnMatch.DoesNotExist:
                selected_column = -1

            attribute_contexts.append(
                {
                    "selected_column": selected_column,
                    "asset_attribute": attr_name,
                    "verbose_name": verbose_name,
                }
            )

        return render(
            request,
            self.template_name,
            {
                "xls_import": xls_import,
                "skip_lines": xls_import.skip_lines,
                "column_choices": column_choices,
                "attributes": attribute_contexts,
            },
        )

    def post(self, request, pk):
        xls_import = get_object_or_404(XlsImport, pk=pk)

        try:
            xls_import.skip_lines = int(request.POST["skip_lines"])
        except (KeyError, ValueError, TypeError):
            xls_import.skip_lines = 0

        xls_import.save()

        for attr_name, verbose_name in list_importable_attributes():
            try:
                selected_column = int(request.POST[attr_name])
                if selected_column < 0:
                    selected_column = None
            except (KeyError, ValueError, TypeError):
                selected_column = None

            column_match, created = xls_import.column_matches.get_or_create(
                asset_attribute=attr_name
            )
            column_match.column_index = selected_column
            column_match.save()

        return redirect("assets:import-xls-preview", pk=xls_import.pk)


@method_decorator(role_required(User.ROLE_ADMIN), name="dispatch")
class ImportXlsPreviewView(generic.View):
    template_name = "assets/import-xls/preview.html"

    def get(self, request, pk):
        xls_import = get_object_or_404(XlsImport, pk=pk)
        xls_asset_file = XlsAssetsFile(xls_import.file)

        assets = xls_asset_file.import_assets(
            xls_import.skip_lines, list(xls_import.column_matches.all())
        )

        preview_headers = []
        for attr_name, verbose_name in list_importable_attributes():
            preview_headers.append(verbose_name.capitalize())

        preview_rows = []
        for asset in assets:
            row = []
            for attr_name, verbose_name in list_importable_attributes():
                if hasattr(asset, f"get_{attr_name}_display"):
                    get_display = getattr(asset, f"get_{attr_name}_display")
                    row.append(get_display())
                else:
                    row.append(getattr(asset, attr_name) or "")
            preview_rows.append(row)

        return render(
            request,
            self.template_name,
            {
                "xls_import": xls_import,
                "preview_headers": preview_headers,
                "preview_rows": preview_rows,
            },
        )

    def post(self, request, pk):
        xls_import = get_object_or_404(XlsImport, pk=pk)
        xls_asset_file = XlsAssetsFile(xls_import.file)

        assets = xls_asset_file.import_assets(
            xls_import.skip_lines, list(xls_import.column_matches.all())
        )

        for asset in assets:
            asset.save()

        return redirect("assets:assets-list", kind_asset="new")


class ExportXlsView(generic.FormView):
    form_class = ExportXlsForm
    template_name = "assets/export-xls.html"

    def make_workbook(self, assets):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Объявления"

        sheet.cell(row=1, column=1, value="№")
        sheet.cell(row=1, column=2, value="Балансодержатель")
        sheet.cell(row=1, column=3, value="Название")
        sheet.cell(row=1, column=4, value="Вид")
        sheet.cell(row=1, column=5, value="Адрес")
        sheet.cell(row=1, column=6, value="Площадь")
        sheet.cell(row=1, column=7, value="Кадастровый номер")
        sheet.cell(row=1, column=8, value="Состояние")
        sheet.cell(row=1, column=9, value="ФИО контактного лица")
        sheet.cell(row=1, column=10, value="Телефон контактного лица")
        sheet.cell(row=1, column=11, value="Email контактного лица")

        for i in range(11):
            sheet.cell(row=2, column=i + 1, value=str(i + 1))

        for asset_ix, asset in enumerate(assets):
            sheet.cell(row=asset_ix + 3, column=1, value=asset_ix + 1)
            sheet.cell(row=asset_ix + 3, column=2, value=asset.balance_holder)
            sheet.cell(row=asset_ix + 3, column=3, value=asset.name)
            sheet.cell(row=asset_ix + 3, column=4, value=asset.get_type_asset_display())
            sheet.cell(row=asset_ix + 3, column=5, value=asset.address)
            sheet.cell(row=asset_ix + 3, column=6, value=asset.square)
            sheet.cell(row=asset_ix + 3, column=7, value=asset.cadastral_number)
            sheet.cell(row=asset_ix + 3, column=8, value=asset.get_state_display())
            sheet.cell(row=asset_ix + 3, column=9, value=asset.full_name_contact_person)
            sheet.cell(row=asset_ix + 3, column=10, value=asset.phone_contact_person)
            sheet.cell(row=asset_ix + 3, column=11, value=asset.email_contact_person)

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            return tmp.read()

    def form_valid(self, form):
        assets = Asset.objects.all()

        if form.cleaned_data["kind"] == ExportXlsForm.Kind.NEW:
            assets = assets.new_assets()
        elif form.cleaned_data["kind"] == ExportXlsForm.Kind.CONST:
            assets = assets.cost_assets()
        elif form.cleaned_data["kind"] == ExportXlsForm.Kind.ARCHIVED:
            assets = assets.archive_assets()

        assets = assets.annotate(
            resolution_count=models.Count(
                "resolution", filter=models.Q(resolution__kind=Resolution.Kind.APPROVED)
            )
        )

        if (
            form.cleaned_data["resolution_status"]
            == ExportXlsForm.ResolutionStatus.WITH_PRETENDENTS
        ):
            assets = assets.filter(resolution_count__gt=0)
        elif (
            form.cleaned_data["resolution_status"]
            == ExportXlsForm.ResolutionStatus.WITHOUT_PRETENDENTS
        ):
            assets = assets.filter(resolution_count=0)

        xls_bytes = self.make_workbook(assets)
        response = HttpResponse(
            xls_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=assets.xlsx"
        return response
